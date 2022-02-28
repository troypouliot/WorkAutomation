import io
import json
import os
import time
import openpyxl as xl
import shutil
from os import path

########################################################################
#                         INSTRUCTIONS                                 #
########################################################################
#
#   1. Using POSTMAN, send a GET request to each printer.
#       (http://192.168.1.70/server/history/list?limit=5000&order=asc)
#       (http://192.168.1.71/server/history/list?limit=5000&order=asc)
#   2. Save the response as a .json file in the directory:
#       "D:\My Drive\3D Printer\Mainsail\Jobs History" using the
#       machine name as the filename (v0.1, v0.2)
#   3. Run this script. It will extract each of the available .json files
#       and save the print jobs history to the history.xlsx file.
#       Then it will rename and move the .json files to the archive dir.
#

machines = ['v0.1', 'v0.2'] # Add additional machines here if need be

data_file_path = 'D:\\My Drive\\3D Printer\\Mainsail\\Jobs History\\'
log = 'D:\\My Drive\\3D Printer\\Mainsail\\Jobs History\\history.xlsx'

for machine in machines:
    with open(log, 'rb') as f:
        in_mem_file = io.BytesIO(f.read())

        wb = xl.load_workbook(in_mem_file)
        ws = wb[machine]
        try:
            with open(path.join(data_file_path, machine + '.json')) as f:
                data = json.load(f)

            for i in data['result']['jobs']:
                if (int(time.strftime('%H', time.gmtime(i['print_duration']))) == 0) and (int(time.strftime('%M', time.gmtime(i['print_duration']))) < 10):
                    pass
                else:
                    ws.cell(column=1, row=ws.max_row+1, value=i['status']) # Print status
                    ws.cell(column=2, row=ws.max_row, value=time.strftime('%d.%m.%Y', time.strptime(time.ctime(i['start_time'])))) # Print date (Day.Month.Year)
                    ws.cell(column=3, row=ws.max_row, value=int(time.strftime('%H', time.gmtime(i['print_duration'])))) # Print duration (hours)
                    ws.cell(column=4, row=ws.max_row, value=int(time.strftime('%M', time.gmtime(i['print_duration'])))) # Print duration (seconds)
                    ws.cell(column=5, row=ws.max_row, value=i['filename']) # Print filename

            # Create the Archive directory if it doesn't exist
            if not path.exists(path.join(data_file_path, 'Archive')):
                os.mkdir(path.join(data_file_path, 'Archive'))

            # Move and rename the .json file to the archive directory
            shutil.move(path.join(data_file_path, machine + '.json'),
                        (path.join(data_file_path, 'Archive', machine + '_' + time.strftime('%Y%m%d', time.strptime(time.ctime())) + '.json')))
        except FileNotFoundError:
            pass
    wb.save(log)
wb.close()
